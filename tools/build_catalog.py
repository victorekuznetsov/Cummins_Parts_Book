#!/usr/bin/env python3
# =====================================================================
# Сборка интерактивного веб-каталога из выгрузки crawler.py / crawl_details.py.
#
#   python tools/build_catalog.py 37292556 [--prices прайс.xlsx]
#
# Каталог рассчитан на несколько двигателей: каждый ESN кладётся отдельно и
# появляется в переключателе в шапке. Создаётся/обновляется:
#   catalog/data/<ESN>.js    — данные двигателя (window.CATALOGS[<ESN>])
#   catalog/engines.js       — список двигателей для переключателя
#   catalog/drawings/<ESN>/  — чертежи узлов
#   catalog/parts/<ESN>/     — фотографии деталей
#   catalog/index.html       — дописываются <script> на файлы данных
# Оболочка (index.html, app.js, styles.css) статическая и не перезаписывается.
# =====================================================================
import sys, io, json, re, shutil, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CAT  = ROOT / "catalog"

# Русские названия систем — как их показывает сайт Cummins
SYSTEM_RU = {
    "AIR INTAKE": "Воздухозаборник",
    "BASE ENGINE": "Базовый двигатель",
    "COMPRESSORS AND PUMPS": "Компрессоры и насосы",
    "COOLING": "Система охлаждения",
    "DRIVES AND MOUNTINGS": "Приводы и крепления",
    "ELECTRICS": "Электрооборудование",
    "EXHAUST": "Система выпуска",
    "FUEL": "Топливная система",
    "LUBRICATION": "Система смазки",
    "MISCELLANEOUS": "Прочее",
    "RATINGS AND CALIBRATIONS": "Номиналы и калибровки",
    "UNCLASSIFIED": "Без системы",
}

IN_MM, LB_KG = 25.4, 0.45359237


def safe(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", str(name)).strip(". ") or "unnamed"


def drawing_file(fname: str) -> str:
    return safe(fname.strip("/").replace("/", "_")) + ".png"


def _num(v):
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(v or ""))
    return float(m.group(0).replace(",", ".")) if m else None


def _conv(value, to):
    """'7.25 in' -> мм, '2.32 lb' -> кг."""
    n = _num(value)
    if n is None:
        return ""
    s = str(value).lower()
    if to == "mm":
        return f"{n * IN_MM if 'in' in s else n:.0f}"
    n = n * LB_KG if "lb" in s else n
    return f"{n:.2f}".rstrip("0").rstrip(".")


def flatten_parts(groups):
    """Плоский список позиций с сохранением уровня вложенности подкомпонентов."""
    out = []

    def walk(node, level):
        d = node.get("data") or {}
        pn = d.get("partNo")
        if pn or d.get("partDesc"):
            out.append({
                "pos":  (d.get("callOut") or "").strip(),
                "no":   (pn or "").strip(),
                "name": (d.get("partDesc") or "").strip(),
                "qty":  (d.get("qty") or "").strip(),
                "dim":  (d.get("dimensions") or "").strip(),
                "rem":  (d.get("remarks") or "").strip(),
                "lvl":  level,
                "img":  "",
            })
        for ch in (node.get("children") or []):
            walk(ch, level + 1)

    for g in (groups or []):
        for p in (g.get("parts") or []):
            walk(p, 0)
    return out


def load_part_cards(src, part_nos):
    """Карточки деталей: атрибуты, замены номеров, где применяется, ракурсы фото."""
    pdir, cards, views = src / "partdetails", {}, set()
    n_sup = 0
    if not pdir.exists():
        return cards, views, 0
    for pn in sorted(part_nos):
        f = pdir / f"{safe(pn)}.json"
        if not f.exists():
            continue
        d = json.loads(f.read_text(encoding="utf-8"))
        attrs = {}
        for block in (d.get("metadata") or []):
            for a in (block.get("attributes") or []):
                for _, v in a.items():
                    nm, val = (v.get("name") or "").strip(), (v.get("value") or "").strip()
                    if nm and val:
                        attrs[nm] = val
        # цепочка замен: sequence 1 — действующий номер, дальше по убыванию — старые
        sup = []
        for s in (d.get("supersession") or []):
            if not isinstance(s, dict) or not s.get("partNo"):
                continue
            sup.append({
                "no":   str(s["partNo"]).strip(),
                "st":   re.sub(r"^\d+-\s*", "",
                               str(s.get("partSscDesc") or s.get("partSsc") or "")).strip(),
                "sell": (s.get("sellable") == "Y"),
                "seq":  int(_num(s.get("sequence")) or 0),
            })
        sup.sort(key=lambda x: -x["seq"])       # от старого номера к новому
        if len(sup) > 1:
            n_sup += 1
        vs = []
        for g in (d.get("graphics") or []):
            fn = g.get("fileName")
            if fn:
                nm = safe(fn.rsplit("/", 1)[-1])
                if (src / "parts" / nm).exists():
                    vs.append(nm); views.add(nm)
        card = {
            "wt":    _conv(attrs.get("Weight"), "kg"),
            "dim":   "×".join(x for x in (_conv(attrs.get("Length"), "mm"),
                                          _conv(attrs.get("Width"), "mm"),
                                          _conv(attrs.get("Height"), "mm")) if x),
            "attrs": attrs,
            "sup":   sup,
            "recon": d.get("reconEquivalent") or "",
            "used":  [{"o": w.get("item"), "n": w.get("itemDesc") or ""}
                      for w in (d.get("whereUsed") or []) if w.get("itemType") == "O"][:60],
            "views": sorted(vs, key=lambda x: (0 if "_iso." in x else 1, x)),
        }
        cards[pn] = {k: v for k, v in card.items() if v not in ("", [], {}, None)}
    return cards, views, n_sup


def load_prices(path):
    """Прайс-лист xlsx: первый столбец с номером детали, ищем цену и аналитики."""
    if not path:
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("!! openpyxl не установлен (pip install openpyxl) — цены пропущены")
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().lower() for c in next(rows)]

    def col(*keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None

    i_no    = col("номер", "артикул", "part")
    i_price = col("цена", "price", "стоим")
    i_name  = col("наимен", "назв", "descr")
    i_grp   = col("группа", "group")
    i_alt   = col("взаимозамен", "аналог", "замена", "alt")
    if i_no is None or i_price is None:
        print(f"!! в прайсе нет колонок номера/цены (заголовки: {header[:8]}) — цены пропущены")
        return {}
    prices = {}
    for r in rows:
        if not r or i_no >= len(r) or not r[i_no]:
            continue
        key = re.sub(r"\s+", "", str(r[i_no])).upper()
        val = r[i_price] if i_price < len(r) else None
        prices[key] = {
            "price": float(val) if isinstance(val, (int, float)) else None,
            "name":  str(r[i_name]).strip() if i_name is not None and i_name < len(r) and r[i_name] else "",
            "group": str(r[i_grp]).strip() if i_grp is not None and i_grp < len(r) and r[i_grp] else "",
            "alt":   str(r[i_alt]).strip() if i_alt is not None and i_alt < len(r) and r[i_alt] else "",
        }
    print(f">>> Прайс-лист: {len(prices)} позиций из {Path(path).name}")
    return prices


def update_registry(catalog, machine="", fleet=None):
    """catalog/engines.js — список двигателей для переключателя в шапке."""
    reg = CAT / "engines.js"
    engines = {}
    if reg.exists():
        m = re.search(r"window\.ENGINES\s*=\s*(\[.*?\]);", reg.read_text(encoding="utf-8"), re.S)
        if m:
            for e in json.loads(m.group(1)):
                engines[e["esn"]] = e
    prev = engines.get(catalog["esn"], {})
    engines[catalog["esn"]] = {
        "esn": catalog["esn"], "model": catalog["model"], "cpl": catalog["cpl"],
        "machine": machine or prev.get("machine", ""),
        "build": catalog["buildDate"], "config": catalog["config"],
        "options": len(catalog["options"]),
        "parts": len({p["no"] for o in catalog["options"] for p in o["parts"] if p["no"]}),
        "fleet": fleet if fleet is not None else prev.get("fleet", []),
    }
    rows = sorted(engines.values(), key=lambda e: (str(e["machine"]), str(e["model"])))
    reg.write_text("window.ENGINES = " + json.dumps(rows, ensure_ascii=False, indent=1) + ";\n",
                   encoding="utf-8")
    return rows


def update_index(engines):
    """Подключаем в index.html engines.js и файлы данных всех двигателей."""
    idx = CAT / "index.html"
    html = idx.read_text(encoding="utf-8")
    tags = ['<script src="engines.js"></script>'] + \
           [f'<script src="data/{e["esn"]}.js"></script>' for e in engines]
    block = "\n".join(tags)
    html = re.sub(r'(?:<script src="(?:engines\.js|data/\d+\.js)"></script>\s*)+', "", html)
    html = html.replace('<script src="app.js"></script>', block + '\n<script src="app.js"></script>')
    idx.write_text(html, encoding="utf-8")


def load_fleet(report_path, esn):
    """Остальные ESN с тем же CPL — из отчёта tools/check_esn.py."""
    if not report_path:
        return None
    d = json.loads(Path(report_path).read_text(encoding="utf-8"))
    for g in d.get("groups", []):
        if esn in g.get("esns", []):
            return [e for e in g["esns"] if e != esn]
    print(f"!! {esn} не найден в {report_path} — список парка не заполнен")
    return None


def build(esn, prices_path=None, machine="", fleet_report=None):
    src = ROOT / "data" / esn
    if not (src / "engine.json").exists():
        sys.exit(f"нет выгрузки {src} — сначала запустите crawler.py {esn}")

    engine = json.loads((src / "engine.json").read_text(encoding="utf-8"))
    prices = load_prices(prices_path)

    # система -> варианты исполнения (берём из состава двигателя)
    sys_of_option = {}
    for o in (engine.get("optionList") or []):
        s = {x for p in (o.get("parts") or []) for x in (p.get("systems") or [])}
        sys_of_option[o.get("optionNo")] = sorted(s) or ["UNCLASSIFIED"]

    options, all_sheets, all_photos = [], set(), set()
    for f in sorted((src / "options").glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        no = d.get("optionNo") or f.stem
        parts = flatten_parts(d.get("groups"))
        sheets = [drawing_file(g["fileName"]) for g in (d.get("graphics") or [])
                  if g.get("fileName") and (src / "drawings" / drawing_file(g["fileName"])).exists()]
        all_sheets.update(sheets)
        for p in parts:
            pr = prices.get(re.sub(r"\s+", "", p["no"]).upper())
            if pr:
                p["price"] = pr["price"]
                if pr["group"]: p["group"] = pr["group"]
                if pr["alt"]:   p["alt"] = pr["alt"]
            photo = f"{safe(p['no'])}_iso.png"
            if p["no"] and (src / "parts" / photo).exists():
                p["img"] = photo
                all_photos.add(photo)
        options.append({
            "no": no,
            "name": d.get("optionName") or no,
            "systems": sys_of_option.get(no, ["UNCLASSIFIED"]),
            "remarks": (d.get("remarks") or "").strip(),
            "sheets": sheets,
            "parts": parts,
        })

    # ремкомплекты
    kits = []
    kf = src / "kitSets.json"
    if kf.exists():
        for k in json.loads(kf.read_text(encoding="utf-8")):
            kits.append({
                "no": k.get("kitNo"), "name": k.get("kitDesc"),
                "notes": k.get("kitNotes") or "", "type": k.get("kitType") or "",
                "parts": [{"no": p.get("partNo"), "name": p.get("partDesc")}
                          for p in (k.get("parts") or [])],
            })

    # дерево систем
    systems = []
    for code in sorted({s for o in options for s in o["systems"]}):
        systems.append({
            "code": code,
            "name": SYSTEM_RU.get(code, code.title()),
            "options": sorted([o["no"] for o in options if code in o["systems"]]),
        })

    uniq_nos = {p["no"] for o in options for p in o["parts"] if p["no"]}
    cards, card_views, n_sup = load_part_cards(src, uniq_nos)
    all_photos.update(card_views)

    catalog = {
        "esn": esn,
        "model": engine.get("serviceModel"),
        "cpl": engine.get("cpl"),
        "buildDate": str(engine.get("buildDate") or "")[:10],
        "config": engine.get("marketingConfig"),
        "group": engine.get("engineGroup"),
        "plant": engine.get("enginePlantCode"),
        "hasPrices": bool(prices),
        "systems": systems,
        "options": options,
        "kits": kits,
        "cards": cards,
    }

    (CAT / "data").mkdir(parents=True, exist_ok=True)
    (CAT / "data" / f"{esn}.js").write_text(
        "window.CATALOGS = window.CATALOGS || {};\nwindow.CATALOGS[\"" + esn + "\"] = " +
        json.dumps(catalog, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8")

    # картинки — в подпапки по ESN
    for sub, names in (("drawings", all_sheets), ("parts", all_photos)):
        dst = CAT / sub / esn
        dst.mkdir(parents=True, exist_ok=True)
        for n in names:
            s = src / sub / n
            if s.exists() and not (dst / n).exists():
                shutil.copy2(s, dst / n)

    engines = update_registry(catalog, machine, load_fleet(fleet_report, esn))
    update_index(engines)

    total_pos = sum(len(o["parts"]) for o in options)
    print(f">>> Двигатель {esn} ({catalog['model']}, CPL {catalog['cpl']}) добавлен в каталог")
    print(f"    систем {len(systems)}, узлов {len(options)}, позиций {total_pos}, "
          f"уникальных деталей {len(uniq_nos)}")
    print(f"    чертежей {len(all_sheets)}, фото деталей {len(all_photos)}, "
          f"ремкомплектов {len(kits)}")
    print(f"    карточек деталей {len(cards)}, из них с заменами номеров {n_sup}")
    print(f"    data/{esn}.js: {(CAT / 'data' / f'{esn}.js').stat().st_size / 1024:.0f} КБ")
    print(f"    всего двигателей в каталоге: {len(engines)} "
          f"({', '.join(e['esn'] for e in engines)})")

    from_engine = {p["partNo"] for o in (engine.get("optionList") or [])
                   for p in (o.get("parts") or []) if p.get("partNo")}
    lost = sorted(from_engine - uniq_nos)
    print(f"    ПОТЕРЯНО НОМЕРОВ: {len(lost)}" + (f" -> {lost[:10]}" if lost else "  (ноль)"))
    return len(lost) == 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Сборка веб-каталога Cummins")
    ap.add_argument("esn")
    ap.add_argument("--prices", help="прайс-лист .xlsx (необязательно)")
    ap.add_argument("--machine", default="", help="машина, на которой стоит двигатель (например, NTE200)")
    ap.add_argument("--fleet-from", dest="fleet", default=None,
                    help="отчёт check_esn.py — подставит остальные ESN того же CPL")
    a = ap.parse_args()
    sys.exit(0 if build(a.esn, a.prices, a.machine, a.fleet) else 1)
